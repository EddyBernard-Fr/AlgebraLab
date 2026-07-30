from PySide6.QtWidgets import (
    QMainWindow,
    QWidget,
    QPushButton,
    QListWidget,
    QVBoxLayout,
    QApplication,
    QTableWidget,
    QSpinBox,
    QLineEdit,
    QLabel,
    QHBoxLayout,
    QGridLayout,
    QTableWidgetItem,
    QMessageBox,
    QFrame,
    QComboBox,
    QFileDialog,  
    QMessageBox
)

from PySide6.QtCore import Qt
from algebralab.manager.utils import *
from algebralab.manager.matrix_manager import *
import numpy as np
import os
from openpyxl import load_workbook




class AccueilPage(QWidget):

    def __init__(self, manager, main_window):
        super().__init__()

        # Paramètres de la fenêtre
        
        self.manager = manager
        self.main_window = main_window

        self.layout = QHBoxLayout()
        self.layout.setSpacing(15)

        self.nom_label = QLabel("Nom :")
        self.nom_input = QLineEdit()
        self.lignes_label = QLabel("Lignes :")
        self.lignes_input = QSpinBox()
        self.colonnes_label = QLabel("Colonnes :")
        self.colonnes_input = QSpinBox()
        self.bouton_creer = QPushButton("Créer grille") 
        
        self.lignes_input.setMinimum(1)
        self.colonnes_input.setMinimum(1)

        self.lignes_input.setValue(3)
        self.colonnes_input.setValue(3)
        self.bouton_creer.setFixedHeight(40)
        self.nom_input.setMinimumWidth(150)
        

        self.liste_matrices = QListWidget()
        self.liste_matrices.setMaximumHeight(200)


        self.zone_liste = QWidget()
        self.liste_layout = QVBoxLayout()

        self.label_liste = QLabel("Matrices disponibles :")

        self.liste_layout.addWidget(
            self.label_liste
        )

        self.liste_layout.setSpacing(30)

        self.liste_layout.addWidget(
            self.liste_matrices
        )

        self.liste_layout.setSpacing(30)

        separateur = QFrame()
        separateur.setFrameShape(QFrame.HLine)
        separateur.setFrameShadow(QFrame.Sunken)

        separateur.setStyleSheet("""
            QFrame {
                color: gray;
            }
        """)

        separateur.setLineWidth(2)

        self.liste_layout.addWidget(separateur)

        self.label_creer = QLabel("Création de matrices:")
        
        self.liste_layout.addWidget(
            self.label_creer
        )

        self.liste_layout.setSpacing(30)

        self.form_layout = QGridLayout()
        self.form_layout.setColumnMinimumWidth(0, 80)
        self.form_layout.setColumnStretch(0, 0)
        self.form_layout.setColumnStretch(1, 1)

        self.form_layout.addWidget(self.nom_label, 0, 0)
        self.form_layout.addWidget(self.nom_input,0,1,alignment=Qt.AlignLeft)

        self.form_layout.addWidget(self.lignes_label, 1, 0)
        self.form_layout.addWidget(self.lignes_input, 1, 1, alignment=Qt.AlignLeft)


        self.form_layout.addWidget(self.colonnes_label, 2, 0)
        self.form_layout.addWidget(self.colonnes_input, 2, 1, alignment=Qt.AlignLeft)

        self.liste_layout.addLayout(self.form_layout)

        self.bouton_creer_layout = QHBoxLayout()

        self.bouton_creer_layout.addStretch()
        self.bouton_creer_layout.addWidget(self.bouton_creer)
        self.bouton_creer_layout.addStretch()

        self.liste_layout.setSpacing(30)

        self.liste_layout.addLayout(
            self.bouton_creer_layout
        )


        self.bouton_layout = QHBoxLayout()

        self.label_import = QLabel("Importer une matrice:")

        self.bouton_import = QPushButton("Importer")

        self.bouton_layout.addWidget(self.label_import)
        self.bouton_layout.addWidget(self.bouton_import)
        self.bouton_layout.addStretch()
        
        
        self.liste_layout.addLayout(self.bouton_layout)




        separateur1 = QFrame()
        separateur1.setFrameShape(QFrame.HLine)
        separateur1.setFrameShadow(QFrame.Sunken)

        separateur1.setStyleSheet("""
            QFrame {
                color: gray;
            }
        """)

        separateur1.setLineWidth(2)

        self.liste_layout.addWidget(separateur1)

        self.label_calculer = QLabel("Opérations:")
        
        self.liste_layout.addWidget(
            self.label_calculer
        )

        self.liste_layout.setSpacing(30)

        self.matrice1_combo = QComboBox()
        self.operation_combo = QComboBox()
        self.matrice2_combo = QComboBox()


        for matrice in self.manager.get_matrices():
            self.matrice1_combo.addItem(matrice.name)
            self.matrice2_combo.addItem(matrice.name)

        self.operation_combo.addItems([
            "+",
            "×"
        ])

        self.form_layout1 = QGridLayout()
        self.form_layout1.setColumnMinimumWidth(0, 80)
        self.form_layout1.setColumnStretch(0, 0)
        self.form_layout1.setColumnStretch(1, 1)

        self.form_layout1.addWidget(QLabel("Matrice 1:"), 0, 0)
        self.form_layout1.addWidget(self.matrice1_combo,0,1,alignment=Qt.AlignLeft)

        self.form_layout1.addWidget(QLabel("Opération :"), 1, 0)
        self.form_layout1.addWidget(self.operation_combo, 1, 1, alignment=Qt.AlignLeft)


        self.form_layout1.addWidget(QLabel("Matrice 2:"), 2, 0)
        self.form_layout1.addWidget(self.matrice2_combo, 2, 1, alignment=Qt.AlignLeft)

        self.liste_layout.addLayout(self.form_layout1)

        self.bouton_calculer_layout = QHBoxLayout()

        self.bouton_calculer = QPushButton("Calculer")

        self.bouton_calculer.setFixedHeight(40)
        

        self.bouton_calculer_layout.addStretch()
        self.bouton_calculer_layout.addWidget(self.bouton_calculer)
        self.bouton_calculer_layout.addStretch()

        self.liste_layout.setSpacing(30)

        self.liste_layout.addLayout(
            self.bouton_calculer_layout
        )


        self.liste_layout.addStretch()

        self.zone_liste.setLayout(
            self.liste_layout
        )

        self.layout.addWidget(
            self.zone_liste
            )
        
        ligne = QFrame()
        ligne.setFrameShape(QFrame.VLine)
        ligne.setFrameShadow(QFrame.Sunken)

        ligne.setStyleSheet("""
            QFrame {
                color: gray;
            }
        """)

        ligne.setLineWidth(2)

        self.layout.addWidget(ligne)

        self.zone_affichage = QWidget()
        self.zone_affichage.setMinimumWidth(500)
        self.layout.addWidget(self.zone_affichage)
        self.layout.setStretch(0, 1)
        self.layout.setStretch(2, 3)

        self.setLayout(self.layout)

        self.table = QTableWidget()
        self.nom_label = QLabel()

        self.bouton_modifier = QPushButton("Modifier")
        self.bouton_supprimer = QPushButton("Supprimer")

        self.bouton_modifier.setFixedHeight(40)
        self.bouton_supprimer.setFixedHeight(40)

        self.bouton_modifier.hide()
        self.bouton_supprimer.hide()

        self.parenthese_gauche = QLabel("(")
        self.parenthese_droite = QLabel(")")


        self.nom_label.hide()
        self.parenthese_gauche.hide()
        self.parenthese_droite.hide()
        self.table.hide()

        self.parenthese_gauche.setAlignment(Qt.AlignCenter)
        self.parenthese_droite.setAlignment(Qt.AlignCenter)


        for matrice in self.manager.get_matrices():
            self.liste_matrices.addItem(matrice.name)


        self.bouton_layout = QHBoxLayout()

        self.bouton_layout.addStretch()

        self.bouton_layout.addWidget(
            self.bouton_modifier
        )

        self.bouton_layout.setSpacing(30)

        self.bouton_layout.addWidget(
            self.bouton_supprimer
        )

        self.bouton_layout.addStretch()




        self.affichage_layout = QHBoxLayout()

        self.affichage_layout.addStretch()


        self.affichage_layout.addWidget(
            self.nom_label
        )

        self.affichage_layout.addWidget(
            self.parenthese_gauche
        )

        self.affichage_layout.addWidget(
            self.table
        )

        self.affichage_layout.addWidget(
             self.parenthese_droite
        )

        self.affichage_layout.addStretch()




        self.layout_affichage = QVBoxLayout()

        self.layout_affichage.addStretch()

        self.layout_affichage.addLayout(
            self.affichage_layout
        )
        
        self.layout_affichage.addSpacing(30)
        

        self.layout_affichage.addLayout(
            self.bouton_layout
        )

        self.save_layout = QHBoxLayout()

        self.bouton_save = QPushButton("Sauvegarder")  
        self.bouton_save.hide()

        self.bouton_save.setFixedHeight(40)

        self.save_layout.addStretch()
        self.save_layout.addWidget(self.bouton_save)
        self.save_layout.addStretch()

        self.layout_affichage.addLayout(
            self.save_layout
        )

        self.layout_affichage.addStretch()

        self.zone_affichage.setLayout(
            self.layout_affichage
        )


        # Widgets
   

        # Connexions
        self.liste_matrices.itemClicked.connect(
            self.afficher_matrice
        )

        self.bouton_supprimer.clicked.connect(
            self.supprimer_matrice
        )

        self.bouton_modifier.clicked.connect(
            self.modifier_matrice
        )

        self.bouton_creer.clicked.connect(self.creer_grille)

        self.bouton_save.clicked.connect(self.sauvegarder_grille)

        self.bouton_calculer.clicked.connect(self.calculer)

        self.bouton_import.clicked.connect(self.import_matrix)


    def afficher_matrice(self, item):


        nom = item.text()
        self.nom_selectionne = nom

        self.table.clear()
        self.table.setShowGrid(False)
        self.table.horizontalHeader().hide()
        self.table.verticalHeader().hide()
        

        matrice = None

        for m in self.manager.get_matrices():
            if m.name == nom:
                matrice = m
                break

        self.nom_label.setText(
            f"{nom} = "
        )

        self.table.setRowCount(matrice.lignes)
        self.table.setColumnCount(matrice.colonnes)
        self.table.setFixedSize(matrice.colonnes * 50, matrice.lignes * 50)

        taille_parenthese = self.table.height()

        for p in [self.parenthese_gauche, self.parenthese_droite]:
            p.setStyleSheet(
            f"font-size: {taille_parenthese}px;"
            )
            p.setFixedHeight(taille_parenthese)
            p.setAlignment(Qt.AlignCenter)

        decalage = -(matrice.lignes * 15)
        self.parenthese_gauche.setContentsMargins(0, decalage, 0, 0)
        self.parenthese_droite.setContentsMargins(0, decalage, 0, 0)

        

        for i in range(matrice.lignes):
            for j in range(matrice.colonnes):

                cellule = QTableWidgetItem(
                  str(matrice.values[i][j])
                )

                cellule.setTextAlignment(
                  Qt.AlignCenter
                )

                self.table.setItem(i,j,cellule) 

        for i in range(matrice.colonnes):
            self.table.setColumnWidth(i, 48)

        for i in range(matrice.lignes):
            self.table.setRowHeight(i, 48)


        self.nom_label.show()
        self.parenthese_gauche.show()
        self.parenthese_droite.show()
        self.table.show() 
        self.bouton_modifier.show()
        self.bouton_supprimer.show()   
        self.bouton_save.hide()  

    def supprimer_matrice(self):

        reponse = QMessageBox.question(
            self,
            "Confirmation",
            "Êtes-vous sûr de vouloir supprimer cette matrice ?",
            QMessageBox.Yes | QMessageBox.No
            )
        
        if reponse == QMessageBox.Yes:
            
            nom = self.nom_selectionne

            if nom is None:
                return

            self.manager.supprimer_matrice_bdd(nom)

            self.liste_matrices.clear()
            self.matrice1_combo.clear()
            self.matrice2_combo.clear()

            for matrice in self.manager.get_matrices():
                self.liste_matrices.addItem(matrice.name)
                self.matrice1_combo.addItem(matrice.name)
                self.matrice2_combo.addItem(matrice.name)

            self.table.clear()

            self.nom_label.hide()
            self.table.hide()
            self.parenthese_gauche.hide()
            self.parenthese_droite.hide()
            
            self.bouton_modifier.hide()
            self.bouton_supprimer.hide()
        else:
            return

    def modifier_matrice(self):
        
        reponse_m = QMessageBox.question(
            self,
            "Confirmation",
            "Êtes-vous sûr de vouloir modifier cette matrice ?",
            QMessageBox.Yes | QMessageBox.No
            )
        
        if reponse_m == QMessageBox.Yes:
            
            nom = self.nom_selectionne

            if nom is None:
                return
            
            lignes = self.table.rowCount()
            colonnes = self.table.columnCount()
            valeurs = []

            for i in range(lignes):

                ligne = []

                for j in range(colonnes):

                    item = self.table.item(i, j)

                    if item is None:
                        ligne.append(0)

                    else:
                        ligne.append(float(item.text()))

                valeurs.append(ligne)

            
            matrice = self.manager.recuperer_matrice(nom)
            matrice.values = valeurs
            update_matrice(nom, valeurs)

            
        else:
            return

    


    def sauvegarder_grille(self):

        nom = self.nom_input.text()

        lignes = self.lignes_input.value()
        colonnes = self.colonnes_input.value()

        valeurs = []

        for i in range(lignes):

            ligne = []

            for j in range(colonnes):

                item = self.table.item(i, j)

                if item is None:
                    ligne.append(0)

                else:
                    try:
                        ligne.append(float(item.text()))
                    except:
                        QMessageBox.warning(
                        self,
                        "Erreur de donnée",
                        "Vous devez entrer des nombres réels."
                        )
                        return

            valeurs.append(ligne)

        
        matrice = objet_matrice(nom, lignes, colonnes, valeurs)
        self.manager.ajouter_matrice_dict(matrice)
        self.manager.save_matrice_bdd(matrice)

        self.liste_matrices.clear()
        self.matrice1_combo.clear()
        self.matrice2_combo.clear()

        for matrice in self.manager.get_matrices():
            self.liste_matrices.addItem(matrice.name)
            self.matrice1_combo.addItem(matrice.name)
            self.matrice2_combo.addItem(matrice.name)



    def creer_grille(self):
        nom = self.nom_input.text()
        lignes = self.lignes_input.value()
        colonnes = self.colonnes_input.value()

        self.nom_label.hide()
        self.bouton_modifier.hide()
        self.bouton_supprimer.hide()
        self.table.clear()

        self.table.setRowCount(lignes)
        self.table.setColumnCount(colonnes)
        self.table.setShowGrid(False)
        self.table.horizontalHeader().hide()
        self.table.verticalHeader().hide()
        self.table.setFixedSize(
            colonnes * 50,
            lignes * 50
        )

        taille_parenthese = self.table.height()

        for p in [self.parenthese_gauche, self.parenthese_droite]:
            p.setStyleSheet(
             f"font-size: {taille_parenthese}px;"
            )
            p.setFixedHeight(taille_parenthese)
            p.setAlignment(Qt.AlignCenter)

        decalage = -(lignes * 15)
        self.parenthese_gauche.setContentsMargins(0, decalage, 0, 0)
        self.parenthese_droite.setContentsMargins(0, decalage, 0, 0)

        for i in range(lignes):
            for j in range(colonnes):
                item = QTableWidgetItem("0")
                item.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(i,j,item)
        
        for i in range(colonnes):
            self.table.setColumnWidth(i, 48)

        for i in range(lignes):
            self.table.setRowHeight(i, 48)

      
        self.parenthese_gauche.show()
        self.table.show()
        self.parenthese_droite.show()
        self.bouton_save.setFixedHeight(40)
        self.bouton_save.show()


        

    def calculer(self):

        self.nom_label.hide()
        self.bouton_modifier.hide()
        self.bouton_supprimer.hide()
        self.bouton_save.hide()
        self.parenthese_gauche.hide()
        self.parenthese_droite.hide()
        self.table.hide()
        self.table.clear()

        nom1 = self.matrice1_combo.currentText()
        operation = self.operation_combo.currentText()
        nom2 = self.matrice2_combo.currentText()
        

        matrice1 = self.manager.recuperer_matrice(nom1)
        matrice2 = self.manager.recuperer_matrice(nom2)

        if operation == "+":

            if matrice1.lignes == matrice2.lignes and matrice1.colonnes == matrice2.colonnes:
                M1 = np.array(matrice1.values)
                M2 = np.array(matrice2.values)

                R = M1 + M2
                if hasattr(R, "tolist"):
                    valeurs = R.tolist()

                matrice = objet_matrice(f"({nom1} + {nom2})", matrice1.lignes, matrice1.colonnes, valeurs)
                self.manager.ajouter_matrice_dict(matrice)
                self.manager.save_matrice_bdd(matrice)

                
                self.table.setRowCount(matrice1.lignes)
                self.table.setColumnCount(matrice1.colonnes)
                self.table.setShowGrid(False)
                self.table.horizontalHeader().hide()
                self.table.verticalHeader().hide()
                self.table.setFixedSize(
                    matrice1.colonnes * 50,
                    matrice1.lignes * 50
                )

                taille_parenthese = self.table.height()

                for p in [self.parenthese_gauche, self.parenthese_droite]:
                    p.setStyleSheet(
                    f"font-size: {taille_parenthese}px;"
                    )
                    p.setFixedHeight(taille_parenthese)
                    p.setAlignment(Qt.AlignCenter)

                decalage = -(matrice1.lignes * 15)
                self.parenthese_gauche.setContentsMargins(0, decalage, 0, 0)
                self.parenthese_droite.setContentsMargins(0, decalage, 0, 0)

                for i in range(matrice1.lignes):
                    for j in range(matrice1.colonnes):

                        cellule = QTableWidgetItem(
                        str(valeurs[i][j])
                        )

                        cellule.setTextAlignment(
                         Qt.AlignCenter
                        )

                        self.table.setItem(i,j,cellule) 
                
                
                
                for i in range(matrice1.colonnes):
                    self.table.setColumnWidth(i, 48)

                for i in range(matrice1.lignes):
                    self.table.setRowHeight(i, 48)


                self.nom_label.setText(
                    f"{nom1} + {nom2} = "
                )
                self.nom_label.show()
                self.parenthese_gauche.show()
                self.table.show()
                self.parenthese_droite.show()

                self.liste_matrices.clear()
                self.matrice1_combo.clear()
                self.matrice2_combo.clear()

                for matrice in self.manager.get_matrices():
                    self.liste_matrices.addItem(matrice.name)
                    self.matrice1_combo.addItem(matrice.name)
                    self.matrice2_combo.addItem(matrice.name)


            else:
                QMessageBox.warning(
                        self,
                        "Erreur de dimension",
                        "" \
                        "Les dimensions des deux matrices ne sont pas compatibles pour cette opération."
                    )
                return
            
        else:
            if matrice1.colonnes == matrice2.lignes:
                M1 = np.array(matrice1.values)
                M2 = np.array(matrice2.values)

                R = np.dot(matrice1.values,matrice2.values)
                if hasattr(R, "tolist"):
                    valeurs = R.tolist()

                matrice = objet_matrice(f"{nom1} x {nom2}", matrice1.lignes, matrice2.colonnes, valeurs)
                self.manager.ajouter_matrice_dict(matrice)
                self.manager.save_matrice_bdd(matrice)

                
                self.table.setRowCount(matrice1.lignes)
                self.table.setColumnCount(matrice2.colonnes)
                self.table.setShowGrid(False)
                self.table.horizontalHeader().hide()
                self.table.verticalHeader().hide()
                self.table.setFixedSize(
                    matrice2.colonnes * 50,
                    matrice1.lignes * 50
                )

                taille_parenthese = self.table.height()

                for p in [self.parenthese_gauche, self.parenthese_droite]:
                    p.setStyleSheet(
                    f"font-size: {taille_parenthese}px;"
                    )
                    p.setFixedHeight(taille_parenthese)
                    p.setAlignment(Qt.AlignCenter)

                decalage = -(matrice1.lignes * 15)
                self.parenthese_gauche.setContentsMargins(0, decalage, 0, 0)
                self.parenthese_droite.setContentsMargins(0, decalage, 0, 0)

                for i in range(matrice1.lignes):
                    for j in range(matrice2.colonnes):

                        cellule = QTableWidgetItem(
                        str(valeurs[i][j])
                        )

                        cellule.setTextAlignment(
                         Qt.AlignCenter
                        )

                        self.table.setItem(i,j,cellule) 
                
                
                
                for i in range(matrice2.colonnes):
                    self.table.setColumnWidth(i, 48)

                for i in range(matrice1.lignes):
                    self.table.setRowHeight(i, 48)


                self.nom_label.setText(
                    f"{nom1} x {nom2} = "
                )
                self.nom_label.show()
                self.parenthese_gauche.show()
                self.table.show()
                self.parenthese_droite.show()

                self.liste_matrices.clear()
                self.matrice1_combo.clear()
                self.matrice2_combo.clear()

                for matrice in self.manager.get_matrices():
                    self.liste_matrices.addItem(matrice.name)
                    self.matrice1_combo.addItem(matrice.name)
                    self.matrice2_combo.addItem(matrice.name)
            else:
                QMessageBox.warning(
                        self,
                        "Erreur de dimension",
                        "" \
                        "Les dimensions des deux matrices ne sont pas compatibles pour cette opération."
                    )
                return
        
                

    def import_matrix(self):
        
        fichier, _ = QFileDialog.getOpenFileName(
            self,
            "Importer une matrice",
            "",
            "Matrices (*.txt *.xlsx)"
        )


        extension = os.path.splitext(fichier)[1].lower()

        if extension == ".txt":
            valeurs = self.importer_txt(fichier)

        elif extension == ".xlsx":
            valeurs = self.importer_excel(fichier)
            if valeurs == None:
                return

        else:
            return
        

        nom = os.path.splitext(
                os.path.basename(fichier)
            )[0]

        lignes = len(valeurs)
        colonnes = len(valeurs[0])

        for ligne in valeurs:

            if len(ligne) != colonnes:
                QMessageBox.warning(
                    self,
                    "Erreur d'import",
                    "Toutes les lignes doivent avoir le même nombre de colonnes."
                )
                return

        matrice = objet_matrice(
            nom,
            lignes,
            colonnes,
            valeurs
        )
        self.manager.ajouter_matrice_dict(matrice)
        self.manager.save_matrice_bdd(matrice)   

        self.liste_matrices.clear()
        self.matrice1_combo.clear()
        self.matrice2_combo.clear()

        for matrice in self.manager.get_matrices():
            self.liste_matrices.addItem(matrice.name)
            self.matrice1_combo.addItem(matrice.name)
            self.matrice2_combo.addItem(matrice.name)

         

    def importer_txt(self, fichier):
        
        if fichier:
            with open(fichier, "r") as f:
                lignes = f.readlines()

            valeurs = []

            for ligne in lignes:
                ligne = ligne.strip()

                if "," in ligne:
                    elements = ligne.split(",")
                else:
                    elements = ligne.split()

                valeurs.append(
                    [float(x) for x in elements]
                )
            return valeurs


    
    def importer_excel(self, fichier):

        wb = load_workbook(fichier, data_only=True)

        ws = wb.active

        valeurs = []

        for row in ws.iter_rows(values_only=True):

            ligne = []

            for cellule in row:

                if cellule is None:
                    QMessageBox.warning(
                    self,
                    "Erreur d'import",
                    "Le fichier contient des cellules vides."
                    )
                    return 

                else:
                    ligne.append(float(cellule))

            valeurs.append(ligne)

        return valeurs

         
